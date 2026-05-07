import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "outputs")


def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_fill():
    return PatternFill("solid", fgColor="D6E4F0")


# ─────────────────────────────────────────────
# 生成发给上游厂商的询价Excel
# ─────────────────────────────────────────────

def gen_upstream_inquiry(items: list[dict], brand: str = "BALLUFF") -> str:
    """
    生成发给厂商的标准询价单Excel。
    格式参考：巴鲁夫报价1_6.xlsx 的厂商格式。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "询价单"

    # 标题
    ws.merge_cells("A1:G1")
    ws["A1"] = f"{brand} 产品询价单"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = f"询价日期：{datetime.now().strftime('%Y年%m月%d日')}"
    ws["A2"].font = Font(size=10)

    # 表头
    headers = ["序号", "完整型号", "订货号", "产品描述", "数量", "单位", "品牌"]
    col_widths = [6, 40, 16, 40, 8, 6, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    # 数据行
    for i, item in enumerate(items, 1):
        row = 4 + i
        values = [
            i,
            item.get("model_full", ""),
            item.get("model_short", ""),
            item.get("description", ""),
            item.get("qty", 0),
            item.get("unit", "个"),
            item.get("brand", brand),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _thin_border()
            cell.alignment = Alignment(
                horizontal="center" if col in [1, 5, 6] else "left",
                wrap_text=True,
                vertical="center"
            )
        ws.row_dimensions[row].height = 30

    # 备注行
    note_row = 4 + len(items) + 2
    ws.cell(row=note_row, column=1, value="备注：请提供含税单价及预计货期，谢谢！")
    ws.cell(row=note_row, column=1).font = Font(italic=True, color="666666")

    filename = f"询价单_{brand}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(path)
    return path


# ─────────────────────────────────────────────
# 生成发给客户的报价Excel
# ─────────────────────────────────────────────

def gen_customer_quote(items: list[dict], client_name: str = "", markup_rate: float = 1.30) -> str:
    """
    生成发给客户的含税报价单Excel。
    格式参考：巴鲁夫-1_报价.xlsx。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "报价单"

    # 表头信息
    ws.merge_cells("A1:H1")
    ws["A1"] = "上海库胜自动化工程有限公司"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = "报  价  单"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws["A3"] = f"客户：{client_name}"
    ws["E3"] = f"日期：{datetime.now().strftime('%Y年%m月%d日')}"
    ws["G3"] = "有效期：30天"
    for cell in [ws["A3"], ws["E3"], ws["G3"]]:
        cell.font = Font(size=10)

    # 表头
    headers = ["序号", "完整型号", "订货号", "产品描述", "单位", "数量", "含税单价", "含税总价"]
    col_widths = [6, 38, 14, 36, 6, 6, 12, 12]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = _header_fill()
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w

    # 数据行
    total_amount = 0.0
    for i, item in enumerate(items, 1):
        row = 5 + i
        sale_price = item.get("sale_price", 0.0)
        qty = item.get("qty", 0)
        total_price = round(sale_price * qty, 2)
        total_amount += total_price

        values = [
            i,
            item.get("model_full", ""),
            item.get("model_short", ""),
            item.get("description", ""),
            item.get("unit", "个"),
            qty,
            sale_price if sale_price > 0 else "待定",
            total_price if sale_price > 0 else "待定",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = _thin_border()
            cell.alignment = Alignment(
                horizontal="center" if col in [1, 5, 6] else "left",
                wrap_text=True,
                vertical="center"
            )
            # 金额列右对齐
            if col in [7, 8] and isinstance(val, float):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0.00"
        ws.row_dimensions[row].height = 30

    # 合计行
    total_row = 5 + len(items) + 1
    ws.merge_cells(f"A{total_row}:G{total_row}")
    ws.cell(row=total_row, column=1, value="含税合计").font = Font(bold=True)
    ws.cell(row=total_row, column=1).alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=total_row, column=8, value=total_amount)
    total_cell.font = Font(bold=True)
    total_cell.number_format = "#,##0.00"
    total_cell.border = _thin_border()

    # 备注
    note_row = total_row + 2
    ws.cell(row=note_row, column=1, value="备注：以上价格均含增值税，运费由我方承担。")
    ws.cell(row=note_row, column=1).font = Font(italic=True, color="666666", size=9)

    filename = f"报价单_{client_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join(OUTPUT_DIR, filename)
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    wb.save(path)
    return path
